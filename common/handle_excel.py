import openpyxl
import random

class Handle_Excel():
    """用来处理表格，可以读取和写入"""

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):
        wb = openpyxl.load_workbook(filename=self.filename)
        sh = wb[self.sheetname]
        res = list(sh.rows)
        title = [i.value for i in res[0]]
        cases = []
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_excel(self, row, column, value):
        wb = openpyxl.load_workbook(filename=self.filename)
        sh = wb[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        wb.save(self.filename)


if __name__ =='__main__':
    excel = Handle_Excel(r'E:\workspace\python0926\datas\datas.xlsx','register')
    # print(excel.read_excel())
    # excel.write_excel(row=10,column=10,value='yes')
